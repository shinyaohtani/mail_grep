from bs4 import BeautifulSoup
from datetime import datetime
from email import policy
from email.header import decode_header
from email.message import Message
from email.parser import BytesParser
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import cast, Any
from typing import Generator
from urllib.parse import quote
import argparse
import csv
import email
import logging
import os
import re
import sys
import traceback
import unicodedata

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from smart_logging import SmartLogging


class MailIdentifiers:
    def __init__(self, message: "MailMessage") -> None:
        self.message_id = message._id()
        self.date_str = message._date()  # 表示用（YYYY-MM-DD HH:MM:SS or 原文）
        self.date_dt = message.date_dt()  # 並べ替え用
        self.link = message.link()
        self.subj = message.subject_str()
        self.from_addr = message.from_str()
        self.to_addr = message.to_str()

    @property
    def excel_link(self) -> str:
        return f'=HYPERLINK("{self.link}","メール")' if self.link else ""


class HitLine:
    def __init__(
        self,
        mail_keys: MailIdentifiers,
        mail_id: int,
        hit_count: int,
        parttype: str,
        line: str,
    ):
        self.mail_keys = mail_keys
        self.mail_id = mail_id
        self.hit_count = hit_count
        self.parttype = parttype
        self.line = line.strip()

    def values(self) -> list[str | int]:
        ret = [
            self.mail_id,
            self.hit_count,
            self.mail_keys.message_id,
            self.mail_keys.excel_link,
            self.mail_keys.date_str,
            self.mail_keys.from_addr,
            self.mail_keys.to_addr,
            self.mail_keys.subj,
            self.parttype,
            self.line,
        ]
        return [MailStringUtils.sanitize_csv_field(v) for v in ret]


class HitReport:
    HEADERS: list[str] = [
        "mail_id",
        "hit_id",
        "message_id",
        "link",
        "Date",
        "From",
        "To",
        "Subject",
        "Matched Part",
        "Matched Line",
    ]

    def __init__(self) -> None:
        super().__init__()
        self.hit_lines: list[HitLine] = []

    def mail_count(self) -> int:
        return sum(1 for row in self.hit_lines if row.hit_count == 1)

    def append_hit_line(self, line: HitLine) -> None:
        self.hit_lines.append(line)

    def sort(self) -> None:
        def _sort_key(row):
            dt = row.mail_keys.date_dt
            return (dt is None, -(dt.timestamp() if dt else 0))

        self.hit_lines.sort(key=_sort_key)

    def store(self, path: Path) -> None:
        if path.suffix.lower() == ".csv":
            self._store_csv(path)
        elif path.suffix.lower() == ".xlsx":
            self._store_xlsx(path)
        else:
            raise ValueError(f"Unsupported file extension: {path.suffix}")

    def _store_csv(self, csv_path: Path) -> None:
        # 4) BOM 付き UTF-8 で保存（Excel 配慮）
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            # 3) 指定順のカラムでヘッダ行を書き込み
            writer.writerow(HitReport.HEADERS)
            for row in self.hit_lines:
                writer.writerow(row.values())
        logging.info(f"[MailGrep] excel {csv_path}")

    def _store_xlsx(self, xlsx_path: Path) -> None:
        """
        XLSXで保存（openpyxl が必要）。
        - リンク列はExcelのHYPERLINK関数を使う（CSVと同じ）
        - 改行はCSVと同様に可視化（⏎）
        """
        wb: WorkbookType = Workbook()
        ws: Worksheet = cast(Worksheet, wb.active)
        ws.title = "results"

        ws.append(HitReport.HEADERS)
        for row in self.hit_lines:
            ws.append(row.values())

        # フィルタ機能を有効化
        ws.auto_filter.ref = ws.dimensions

        # 列幅自動調整
        try:
            for i, col in enumerate(
                ws.iter_cols(
                    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ),
                1,
            ):
                max_length: int = 0
                for cell in col:
                    cell = cast(Cell, cell)
                    try:
                        cell_value: str = (
                            str(cell.value) if cell.value is not None else ""
                        )
                        max_length = max(max_length, len(cell_value))
                    except Exception:
                        pass
                col_letter: str = get_column_letter(i)
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
        except Exception:
            logging.warning("[MailGrep] 列幅の自動調整に失敗しました。")
            pass

        # hit_id=1のみが表示されるようにデフォルトフィルタを設定（Excelで開いたとき）
        try:
            from openpyxl.worksheet.filters import (
                CustomFilter,
                CustomFilters,
                FilterColumn,
            )

            # hit_id列は2列目（index=1）
            filter_col = FilterColumn(
                colId=1,
                customFilters=CustomFilters(
                    customFilter=[CustomFilter(operator="equal", val="1")]
                ),
            )
            ws.auto_filter.filterColumn = [filter_col]
        except Exception:
            logging.warning("[MailGrep] デフォルトフィルタの設定に失敗しました。")
            pass
        wb.save(xlsx_path)
        logging.info(f"[MailGrep] excel {xlsx_path}")


class MailFolder:
    def __init__(self, root_dir: Path):
        self._root_dir = root_dir

    def mail_paths(self) -> list[Path]:
        return list(self._root_dir.rglob("*.emlx"))


class MailStringUtils:
    @staticmethod
    def sanitize_csv_field(value) -> str:
        if value is None:
            return ""
        s = str(value)
        return s.replace("\r", "").replace("\n", "⏎")

    @staticmethod
    def _read_emlx(path: Path) -> bytes:
        raw = path.read_bytes()
        return raw[raw.find(b"\n") + 1 :] if raw[0:1].isdigit() else raw

    @staticmethod
    def parse_message(mail_path: Path) -> Message:
        """
        .emlx の長さ行をスキップしたあと、
        ヘッダー＋本文をそのまま BytesParser に渡す
        """
        raw_bytes = MailStringUtils._read_emlx(mail_path)
        # 1) 長さ行があればスキップ
        data = raw_bytes
        if data[:1].isdigit():
            nl = data.find(b"\n")
            if nl != -1:
                data = data[nl + 1 :]

        # 2) バイト列を丸ごとパース（ヘッダーも本文もそのまま）
        parser = BytesParser(policy=policy.default)
        try:
            return parser.parsebytes(data)
        except Exception:
            # 万一エラーが起きたらフォールバック
            return email.message_from_bytes(data, policy=policy.default)

    @staticmethod
    def stringify(v: str | bytes | None) -> str:
        if v is None:
            return ""
        if hasattr(v, "addresses"):
            try:
                return str(v)
            except Exception:
                return ""
        if isinstance(v, bytes):
            try:
                return v.decode("utf-8", errors="replace")
            except Exception:
                return ""
        if not isinstance(v, str):
            return str(v)
        return v

    @staticmethod
    def decode_header(value: str | None) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            value = MailStringUtils.remove_crlf(value)
        try:
            parts = decode_header(value)
        except Exception:
            return ""
        out = []
        for text, enc in parts:
            if isinstance(text, bytes):
                status: str = "noErr"
                if enc:
                    try:
                        return_text = text.decode(enc, errors="strict")
                        out.append(return_text)
                        continue
                    except Exception as e:
                        status = f"decode_header: {enc=} decode failed ({e}), trying utf-8 ..."
                out.append(MailStringUtils._decode_header_fallback(text, status))
            else:
                out.append(str(text))
        return "".join(out)

    @staticmethod
    def _decode_header_fallback(text: bytes, status: str) -> str:
        try:
            return text.decode("utf-8", errors="strict")
        except Exception:
            pass
        try:
            return text.decode("latin1", errors="replace")
        except Exception:
            pass
        logging.warning(f"[MailGrep] {status}")
        logging.warning(
            f"[MailGrep] decode_header: could not decode {repr(text[:40])}, outputting raw bytes."
        )
        return repr(text)

    @staticmethod
    def remove_crlf(value: str) -> str:
        if not value:
            return ""
        return value.replace("\r", "").replace("\n", " ")

    @staticmethod
    def _sanitize_header_section(header_str: str) -> str:
        sanitized = []
        prev_colon = False
        for line in header_str.splitlines():
            if ":" in line:
                key, val = line.split(":", 1)
                val = val.replace("\r", " ").replace("\n", " ")
                sanitized.append(f"{key}:{val}")
                prev_colon = True
            else:
                if prev_colon:
                    line = line.replace("\r", " ").replace("\n", " ")
                sanitized.append(line)
                prev_colon = False
        return "\n".join(sanitized)

    @staticmethod
    def _decode_bytes(raw: bytes) -> str:
        return raw.decode("utf-8", errors="ignore")

    @staticmethod
    def _headers_section(raw_str: str) -> str:
        lines = []
        for line in raw_str.splitlines():
            if line.strip() == "":
                break
            lines.append(line)
        return "\n".join(lines)


class MailMessage:
    def __init__(self, mail_path: Path):
        self._msg: Message = MailStringUtils.parse_message(mail_path)
        self._key_strings = MailIdentifiers(self)

    def key_strings(self) -> MailIdentifiers:
        return self._key_strings

    def _id(self) -> str:
        try:
            v = self._msg.get("Message-ID")
            v = MailStringUtils.stringify(v)
            return MailStringUtils.decode_header(v) if v else ""
        except Exception:
            return ""

    def _date(self) -> str:
        try:
            v = self._msg.get("Date")
            v = MailStringUtils.stringify(v)
            if not v:
                return ""
            dt = parsedate_to_datetime(v)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                v = self._msg.get("Date")
                v = MailStringUtils.stringify(v)
                return MailStringUtils.decode_header(v or "")
            except Exception:
                return ""

    def header_lines(self) -> list[str]:
        result = []
        for k in ("Subject", "From", "To", "Date"):
            try:
                v = self._msg.get(k)
                v = MailStringUtils.stringify(v)
                if v is None:
                    continue
                decoded = MailStringUtils.decode_header(v)
                decoded = decoded.replace("\r", "").replace("\n", " ")
                result.append(f"{k}: {decoded}")
            except Exception as e:
                logging.warning(f"[MailGrep] Could not parse {k}: {e}")
                result.append(f"{k}: [INVALID HEADER]")
        return result

    def body_lines(self) -> list[tuple[str, str]]:
        result = []
        for part in self._iter_text_parts():
            payload = part.get_payload(decode=True)
            if not payload:
                continue

            charset: str = part.get_content_charset() or "utf-8"
            html: str
            try:
                html = bytes(payload).decode(charset, errors="replace")
            except Exception:
                html = bytes(payload).decode("utf-8", errors="replace")

            ctype: str = part.get_content_type()
            if ctype == "text/html":
                # ① 生HTMLをそのまま
                result.append((html, "text/html"))

                # ② BeautifulSoupでパース
                soup = BeautifulSoup(html, "html.parser")
                #    不要タグをまるごと削除
                for tag in soup(["head", "script", "style", "meta", "title", "link"]):
                    tag.decompose()

                # ③ テキストのみ改行区切りで取得
                text_only: str = soup.get_text(separator="\n", strip=True)
                for line in text_only.splitlines():
                    if line.strip():
                        result.append((line, "text/html_textonly"))

                # ④ さらに全文連結版も追加
                concat: str = "".join(soup.stripped_strings)
                if concat:
                    result.append((concat, "text/html_concat"))

            else:
                # text/plain
                text: str = html
                for line in text.splitlines():
                    if line.strip():
                        result.append((line, ctype))

        return result

    def date_dt(self) -> datetime | None:
        """並べ替え用に Date を datetime で返す（失敗時は None）"""
        try:
            v = self._msg.get("Date")
            v = MailStringUtils.stringify(v)
            if not v:
                return None
            return parsedate_to_datetime(v)
        except Exception:
            return None

    def link(self) -> str:
        """
        Mail.app で開けるリンクを生成。
        形式: message:%3Cmessage-id%3E  （< と > は URL エンコード）
        """
        message_id = self._id()
        if not message_id:
            return ""
        mid = message_id.strip()
        if not mid.startswith("<"):
            mid = f"<{mid}>"
        # 角括弧や@を含めて完全にエンコード（Excel でもクリック可能）
        return f"message:{quote(mid, safe='')}"

    def subject_str(self) -> str:
        subj = MailStringUtils.decode_header(
            MailStringUtils.stringify(self._msg.get("Subject"))
        )
        return subj

    def from_str(self) -> str:
        from_ = MailStringUtils.decode_header(
            MailStringUtils.stringify(self._msg.get("From"))
        )
        return from_

    def to_str(self) -> str:
        to_ = MailStringUtils.decode_header(
            MailStringUtils.stringify(self._msg.get("To"))
        )
        return to_

    def _iter_text_parts(self) -> Generator[Message, None, None]:
        if hasattr(self._msg, "is_multipart") and self._msg.is_multipart():
            for p in self._msg.walk():
                if p.get_content_type().startswith("text/"):
                    yield p
        else:
            yield self._msg


class SearchPattern:
    def __init__(self, egrep_pattern: str, flags=0):
        self._egrep_pattern = egrep_pattern
        python_pattern = self._egrep_to_python_regex(self._egrep_pattern)
        self._pattern = re.compile(python_pattern, flags | re.DOTALL)

    # --- egrep to Python正規表現（必要なら改良して下さい） ---
    @staticmethod
    def _egrep_to_python_regex(pattern: str) -> str:
        posix_map = {
            r"\[[:digit:]\]": r"\d",
            r"\[[:space:]\]": r"\s",
            r"\[[:alnum:]\]": r"[A-Za-z0-9]",
            r"\[[:alpha:]\]": r"[A-Za-z]",
            r"\[[:lower:]\]": r"[a-z]",
            r"\[[:upper:]\]": r"[A-Z]",
            r"\[[:punct:]\]": r'[!"#$%&\'()*+,\-./:;<=>?@[\\\]^_`{|}~]',
            r"$begin:math:display$[:blank:]$end:math:display$": r"[ \t]",
            r"$begin:math:display$[:xdigit:]$end:math:display$": r"[A-Fa-f0-9]",
            r"$begin:math:display$[:cntrl:]$end:math:display$": r"[\x00-\x1F\x7F]",
            r"$begin:math:display$[:print:]$end:math:display$": r"[ -~]",
            r"$begin:math:display$[:graph:]$end:math:display$": r"[!-~]",
        }
        for k, v in posix_map.items():
            pattern = re.sub(k, lambda m, v=v: v, pattern)
        return pattern

    def match_mail(self, message: MailMessage) -> list[tuple]:
        matches: list[tuple[str, str]] = []
        # 1) ヘッダー行はこれまでどおり検索
        for line in message.header_lines():
            if self._pattern.search(line):
                matches.append(("header", line))

        # 2) 本文行は text/plain と text/html_textonly のみ対象
        for line, parttype in message.body_lines():
            if parttype not in ("text/plain", "text/html_textonly"):
                continue
            if self._pattern.search(line):
                matches.append((parttype, line))

        return matches

    def unique_name(self) -> str:
        """
        検索文字列からデフォルト保存先:
        <cleaned_head16>_<YYYY-MM-DD_HHMMSS>
        - 空白は "_"、正規表現の特殊文字は除去
        - 先頭の非特殊文字のみから最大16文字
        """
        # NFKC正規化 → 空白を "_" に
        s = unicodedata.normalize("NFKC", self._egrep_pattern).replace(" ", "_")
        # アルファベット・数字・アンダースコア以外を除去
        s = re.sub(r"[^\w]", "", s)
        # 連続 "_" の圧縮＆前後の "_" を除去
        s = re.sub(r"_+", "_", s).strip("_")
        # 先頭16文字（空なら "search"）
        head = (s[:16] or "search").lower()
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        return f"{head}_{ts}"


class MailGrepApp:
    def __init__(
        self,
        mail_storage: MailFolder,
        pattern: SearchPattern,
        output_path: Path | None,
    ):
        self._storage = mail_storage
        self._pattern = pattern
        if output_path is None:
            unique_name = pattern.unique_name()
            out_dir = Path("results")
            out_dir.mkdir(parents=True, exist_ok=True)
            output_path = out_dir / f"{unique_name}.csv"
        self._output_path = output_path

    def run(self):
        report: HitReport = HitReport()
        try:
            for mail_id, path in enumerate(self._storage.mail_paths(), 1):
                try:
                    message: MailMessage = MailMessage(path)
                    mail_keys: MailIdentifiers = message.key_strings()
                    for hit_count, (parttype, line) in enumerate(
                        self._pattern.match_mail(message), 1
                    ):
                        if hit_count == 1:
                            print(
                                f"✓ [{parttype}] hit! ({mail_keys.date_str}) {self.line_preview(line)}"
                            )
                        else:
                            print(f"✓ [{parttype}] {self.line_preview(line)}")
                        hit = HitLine(
                            mail_keys,  # use date_dt as a sort key (datetime or None)
                            mail_id,  # mail_id
                            hit_count,  # hit_id
                            parttype,  # Matched Part
                            line.strip(),  # Matched Line
                        )
                        report.append_hit_line(hit)
                except Exception as e:
                    logging.warning(f"[MailGrep] Skipped {path}: {e}")
        except KeyboardInterrupt:
            print()
            logging.warning("[MailGrep]中断されました。ここまでの結果を保存します。")
        finally:
            # 1) Date 降順でソート（None は末尾）
            report.sort()

            # 拡張子が何であろうと2種類とも保存する
            report.store(self._output_path.with_suffix(".xlsx"))
            report.store(self._output_path.with_suffix(".csv"))

            # 3) メール件数をログに表示（hit_id == 1 の行のみカウント）
            logging.info(f"[MailGrep] {report.mail_count()}件を保存しました。")

    @staticmethod
    def line_preview(line: str) -> str:
        max_len = 50
        preview = line.strip()
        if len(preview) > max_len:
            preview = preview[:max_len] + "..."
        return preview


class AppArguments:
    def __init__(self):
        self.parser: argparse.ArgumentParser = self._create_parser()
        self.args: argparse.Namespace | None = None

    def parse(self) -> argparse.Namespace:
        self.args = self.parser.parse_args()
        if self.args is None:
            raise ValueError("Failed to parse arguments")
        return self.args

    @staticmethod
    def _create_parser() -> argparse.ArgumentParser:
        parser = argparse.ArgumentParser(
            description="egrep風にemlxメールをgrepし、CSVに出力するツール"
        )
        parser.add_argument(
            "pattern", metavar="PATTERN", help="検索したい正規表現（egrep互換）"
        )
        parser.add_argument(
            "-i", "--ignore-case", action="store_true", help="大文字・小文字を無視する"
        )
        parser.add_argument(
            "-o",
            "--output",
            type=Path,
            default=None,
            help="出力CSVファイル名（未指定時は自動生成: results/<pattern先頭16>_タイムスタンプ.csv）",
        )
        parser.add_argument(
            "-s",
            "--source",
            type=Path,
            default=Path.home() / "Library" / "Mail" / "V10",
            help="emlxファイルの格納ディレクトリ",
        )
        return parser


def main():
    app_args = AppArguments()
    args = app_args.parse()

    storage = MailFolder(args.source)

    flags = re.IGNORECASE if args.ignore_case else 0
    pattern = SearchPattern(args.pattern, flags)

    app = MailGrepApp(storage, pattern, args.output)
    app.run()


if __name__ == "__main__":
    MAIL_GREP_DEBUG = os.getenv("MAIL_GREP_DEBUG", "0") == "1"
    log_level = logging.DEBUG if MAIL_GREP_DEBUG else logging.INFO
    with SmartLogging(log_level) as env:
        env.set_stream_filter(True)
        try:
            main()
        except Exception as e:
            logging.error(f"[MailGrep] Unhandled exception: {e}")
            logging.error(
                "Unhandled exception at top-level:\n" + traceback.format_exc()
            )
            print("==== FATAL TRACEBACK ====")
            traceback.print_exc(file=sys.stdout)
            exit(1)
