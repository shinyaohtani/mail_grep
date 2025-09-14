from pathlib import Path
from typing import cast
import csv
import logging

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from hit_line import HitLine


class HitReport:
    HEADERS: list[str] = [
        "mail_id",
        "hit_id",
        "message_id",
        "link",
        "Date",
        "From",
        "To",
        "Subject",
        "Matched Part",
        "Matched Line",
    ]

    def __init__(self) -> None:
        super().__init__()
        self.hit_lines: list[HitLine] = []

    def mail_count(self) -> int:
        return sum(1 for row in self.hit_lines if row.hit_count == 1)

    def append_hit_line(self, line: HitLine) -> None:
        self.hit_lines.append(line)

    def sort(self) -> None:
        def _sort_key(row):
            dt = row.mail_keys.date_dt
            return (dt is None, -(dt.timestamp() if dt else 0))

        self.hit_lines.sort(key=_sort_key)

    def store(self, path: Path) -> None:
        if path.suffix.lower() == ".csv":
            self._store_csv(path)
        elif path.suffix.lower() == ".xlsx":
            self._store_xlsx(path)
        else:
            raise ValueError(f"Unsupported file extension: {path.suffix}")

    def _store_csv(self, csv_path: Path) -> None:
        # 4) BOM 付き UTF-8 で保存（Excel 配慮）
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            # 3) 指定順のカラムでヘッダ行を書き込み
            writer.writerow(HitReport.HEADERS)
            for row in self.hit_lines:
                writer.writerow(row.values())
        logging.info(f"[MailGrep] excel {csv_path}")

    def _store_xlsx(self, xlsx_path: Path) -> None:
        """
        XLSXで保存（openpyxl が必要）。
        - リンク列はExcelのHYPERLINK関数を使う（CSVと同じ）
        - 改行はCSVと同様に可視化（⏎）
        """
        wb: WorkbookType = Workbook()
        ws: Worksheet = cast(Worksheet, wb.active)
        ws.title = "results"

        ws.append(HitReport.HEADERS)
        for row in self.hit_lines:
            ws.append(row.values())

        # フィルタ機能を有効化
        ws.auto_filter.ref = ws.dimensions

        # 列幅自動調整
        try:
            for i, col in enumerate(
                ws.iter_cols(
                    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ),
                1,
            ):
                max_length: int = 0
                for cell in col:
                    cell = cast(Cell, cell)
                    try:
                        cell_value: str = (
                            str(cell.value) if cell.value is not None else ""
                        )
                        max_length = max(max_length, len(cell_value))
                    except Exception:
                        pass
                col_letter: str = get_column_letter(i)
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
        except Exception:
            logging.warning("[MailGrep] 列幅の自動調整に失敗しました。")
            pass

        # hit_id=1のみが表示されるようにデフォルトフィルタを設定（Excelで開いたとき）
        try:
            from openpyxl.worksheet.filters import (
                CustomFilter,
                CustomFilters,
                FilterColumn,
            )

            # hit_id列は2列目（index=1）
            filter_col = FilterColumn(
                colId=1,
                customFilters=CustomFilters(
                    customFilter=[CustomFilter(operator="equal", val="1")]
                ),
            )
            ws.auto_filter.filterColumn = [filter_col]
        except Exception:
            logging.warning("[MailGrep] デフォルトフィルタの設定に失敗しました。")
            pass
        wb.save(xlsx_path)
        logging.info(f"[MailGrep] excel {xlsx_path}")
